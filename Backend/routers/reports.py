from datetime import datetime, timezone, timedelta
from typing import Any, List
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook

try:
    from ..database import get_alerts_collection, get_reports_collection, get_db, get_users_collection
    from ..utils.guard_whatsapp_text import normalize_guard_language
except ImportError:
    from database import get_alerts_collection, get_reports_collection, get_db, get_users_collection
    from utils.guard_whatsapp_text import normalize_guard_language


router = APIRouter(prefix="/reports", tags=["reports"])

BASE_DIR = Path(__file__).resolve().parent.parent
CAPTURES_DIR = BASE_DIR / "captures"
EXPORTS_DIR = BASE_DIR / "exports"

ALERT_STATUS_PENDING = "pending"
ALERT_STATUS_CONFIRMED = "confirmed"
ALERT_STATUS_DISMISSED = "dismissed"
ALERT_STATUS_RESOLVED = "resolved"
ALERT_ALLOWED_STATUSES = {
    ALERT_STATUS_PENDING,
    ALERT_STATUS_CONFIRMED,
    ALERT_STATUS_DISMISSED,
    ALERT_STATUS_RESOLVED,
}


def _directory_stats(path: Path) -> tuple[int, int]:
    """Return (file_count, total_bytes) for a directory tree."""
    if not path.exists() or not path.is_dir():
        return 0, 0

    count = 0
    total_bytes = 0
    for file_path in path.rglob("*"):
        if not file_path.is_file():
            continue
        count += 1
        try:
            total_bytes += file_path.stat().st_size
        except OSError:
            continue
    return count, total_bytes


def _coerce_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        try:
            normalized = value.replace("Z", "+00:00")
            dt = datetime.fromisoformat(normalized)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None
    return None


def _normalize_alert_status(value: Any) -> str:
    status = str(value or ALERT_STATUS_PENDING).strip().lower()
    if status == "verified":
        return ALERT_STATUS_RESOLVED
    if status == "false_alarm":
        return ALERT_STATUS_DISMISSED
    if status in ALERT_ALLOWED_STATUSES:
        return status
    return ALERT_STATUS_PENDING


def _safe_ratio(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100.0, 2)


def _normalize_phone_digits(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _camera_location_from_id(camera_id: int | None) -> str | None:
    if camera_id is None:
        return None
    if int(camera_id) == 0:
        return "Camera 1 Area"
    if int(camera_id) == 1:
        return "Camera 2 Area"
    return None


def _resolve_confirmed_frame_path(value: Any) -> str | None:
    """Validate a relative capture path and keep it inside captures/ only."""
    raw = str(value or "").strip()
    if not raw:
        return None

    normalized = raw.replace("\\", "/").lstrip("/")
    parts = Path(normalized).parts
    if not parts or ".." in parts:
        return None

    base = CAPTURES_DIR.resolve()
    candidate = (CAPTURES_DIR / normalized).resolve()
    try:
        candidate.relative_to(base)
    except ValueError:
        return None

    if not candidate.is_file():
        return None
    return normalized


def _time_filter(date_from: str | None = None, date_to: str | None = None) -> dict:
    ts_filter: dict[str, datetime] = {}
    start = _coerce_datetime(date_from) if date_from else None
    end = _coerce_datetime(date_to) if date_to else None
    if start is not None:
        ts_filter["$gte"] = start
    if end is not None:
        ts_filter["$lte"] = end
    return {"timestamp": ts_filter} if ts_filter else {}


def _extract_resolution(alert: dict) -> tuple[str | None, datetime | None, str | None]:
    history = alert.get("action_history") or []
    if not isinstance(history, list):
        return None, None, None

    chosen: tuple[str | None, datetime | None, str | None] = (None, None, None)
    for action in history:
        action_name = _normalize_alert_status(action.get("action"))
        if action_name not in (ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED, ALERT_STATUS_DISMISSED):
            continue
        action_ts = _coerce_datetime(action.get("timestamp"))
        if action_ts is None:
            continue
        actor = action.get("by")
        if chosen[1] is None or action_ts < chosen[1]:
            chosen = (action_name, action_ts, actor)
    return chosen


def _bucket_label(ts: datetime, granularity: str) -> str:
    if granularity == "hourly":
        return ts.strftime("%Y-%m-%d %H:00")
    if granularity == "weekly":
        start = ts - timedelta(days=ts.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        return start.strftime("%Y-%m-%d")
    return ts.strftime("%Y-%m-%d")


def _serialize_alert(doc: dict) -> dict:
    serialized = dict(doc)
    serialized["id"] = str(serialized.pop("_id"))
    serialized.setdefault("subtype", None)
    serialized.setdefault("frame_id", None)
    serialized.setdefault("frame_path", None)
    serialized.setdefault("verified", False)
    serialized["status"] = _normalize_alert_status(serialized.get("status"))
    serialized.setdefault("dispatched_at", None)
    serialized.setdefault("action_history", [])
    serialized.setdefault("location", "AI Camera")
    serialized.setdefault("source_camera_id", None)
    serialized.setdefault("primary_camera_id", None)
    serialized.setdefault("multi_angle_verified", False)
    serialized.setdefault("evidence_urls", [])
    serialized.setdefault("ai_summary_en", None)
    serialized.setdefault("ai_summary_hi", None)
    serialized.setdefault("ai_summary_mr", None)
    serialized.setdefault("ai_narrative_en", None)
    serialized.setdefault("ai_narrative_hi", None)
    serialized.setdefault("ai_narrative_mr", None)
    serialized.setdefault("movement_direction", None)
    serialized.setdefault("movement_confidence", None)
    serialized.setdefault("narrative_generation_mode", None)
    return serialized


def _to_excel_cell_datetime(value: Any) -> str:
    dt = _coerce_datetime(value)
    if dt is None:
        return ""
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


@router.get("/alerts")
async def list_alerts(limit: int = 50):
    """Fetch pending live alerts (real-time detections) sorted by timestamp."""
    alerts_collection = get_alerts_collection()
    # Only return alerts that are still pending (not resolved or marked false alarm)
    cursor = alerts_collection.find({"status": ALERT_STATUS_PENDING}).sort(
        "timestamp", -1).limit(limit)
    alerts: List[dict] = []
    async for doc in cursor:
        alerts.append(_serialize_alert(doc))
    return alerts


@router.get("/verified")
async def list_verified_alerts(limit: int = 50, phone_number: str | None = None, created_after: str | None = None):
    """Fetch guard feed alerts: confirmed only, optionally bounded by guard created_at."""
    alerts_collection = get_alerts_collection()

    query: dict[str, Any] = {"status": ALERT_STATUS_CONFIRMED}
    floor_ts = _coerce_datetime(created_after) if created_after else None

    if phone_number:
        users_collection = get_users_collection()
        normalized_phone = _normalize_phone_digits(phone_number)
        guard = await users_collection.find_one(
            {"role": "guard", "phone_normalized": normalized_phone},
            {"created_at": 1},
        )
        if guard:
            created_at = _coerce_datetime(guard.get("created_at"))
            if created_at and (floor_ts is None or created_at > floor_ts):
                floor_ts = created_at

    if floor_ts is not None:
        query["timestamp"] = {"$gte": floor_ts}

    cursor = alerts_collection.find(query).sort("timestamp", -1).limit(limit)

    alerts: List[dict] = []
    async for doc in cursor:
        alerts.append(_serialize_alert(doc))
    return alerts


@router.get("/alerts/by-type")
async def list_alerts_by_type(alert_type: str = None, limit: int = 50):
    """Fetch alerts grouped by type (weapon, violence, fire)."""
    alerts_collection = get_alerts_collection()
    query = {"status": {
        "$in": [ALERT_STATUS_CONFIRMED, ALERT_STATUS_RESOLVED]}}
    if alert_type:
        query["type"] = alert_type

    cursor = alerts_collection.find(query).sort("timestamp", -1).limit(limit)
    alerts: List[dict] = []
    async for doc in cursor:
        alerts.append(_serialize_alert(doc))
    return alerts


@router.get("/alerts/false-alarms")
async def list_false_alarms(limit: int = 50):
    """Fetch all false alarms."""
    alerts_collection = get_alerts_collection()
    cursor = alerts_collection.find({"status": ALERT_STATUS_DISMISSED}).sort(
        "timestamp", -1).limit(limit)
    alerts: List[dict] = []
    async for doc in cursor:
        alerts.append(_serialize_alert(doc))
    return alerts


@router.patch("/alerts/{alert_id}/verify")
async def verify_alert(alert_id: str, email: str | None = None):
    """Mark an alert as resolved without notifying guards."""
    from bson import ObjectId
    from datetime import datetime

    alerts_collection = get_alerts_collection()
    alert = await alerts_collection.find_one({"_id": ObjectId(alert_id)})
    if not alert:
        return {"modified_count": 0}

    action = {
        "action": ALERT_STATUS_RESOLVED,
        "by": email,
        "timestamp": datetime.now(timezone.utc),
    }

    result = await alerts_collection.update_one(
        {"_id": ObjectId(alert_id)},
        {
            "$set": {"verified": True, "status": ALERT_STATUS_RESOLVED},
            "$push": {"action_history": action},
        },
    )

    # If we know which guard is handling this incident, associate the alert with their active duty record.
    if email:
        try:
            duty_col = get_db()["guard_duty"]
            active = await duty_col.find_one({"email": email, "logout_time": None})
            if active:
                await duty_col.update_one(
                    {"_id": active["_id"]},
                    {"$addToSet": {"alerts_handled": str(alert_id)}},
                )
        except Exception as exc:
            print(f"[REPORTS] Failed to log alert to duty record: {exc}")

    return {"modified_count": result.modified_count}


@router.patch("/alerts/{alert_id}/confirm")
async def confirm_alert(
    alert_id: str,
    email: str | None = None,
    confirmed_frame_path: str | None = None,
    primary_camera_id: int | None = None,
):
    """Mark an alert as a confirmed threat and send notifications."""
    from bson import ObjectId
    from datetime import datetime

    alerts_collection = get_alerts_collection()
    alert_object_id = ObjectId(alert_id)
    alert = await alerts_collection.find_one({"_id": alert_object_id})
    if not alert:
        return {"modified_count": 0}

    alert_status = str(alert.get("status") or "").strip().lower()
    already_dispatched = (
        bool(alert.get("dispatched"))
        or bool(alert.get("dispatched_at"))
        or alert_status in {ALERT_STATUS_CONFIRMED, "dispatched"}
    )
    if already_dispatched:
        return {
            "modified_count": 0,
            "message": "Alert already dispatched.",
            "already_dispatched": True,
            "dispatched_recipients": [],
        }

    # Prefer a fresh confirm-time frame from UI when available.
    selected_frame_path = (
        _resolve_confirmed_frame_path(confirmed_frame_path)
        or _resolve_confirmed_frame_path(alert.get("frame_path"))
    )
    location_value = str(alert.get("location") or "AI Camera")
    if bool(alert.get("multi_angle_verified")) and primary_camera_id is not None:
        override_location = _camera_location_from_id(primary_camera_id)
        if override_location:
            location_value = override_location
    incident_time_value = (
        alert.get("timestamp").isoformat()
        if alert.get("timestamp")
        else datetime.now(timezone.utc).isoformat()
    )

    narrative_payload: dict[str, Any] = {
        "ai_summary_en": None,
        "ai_summary_hi": None,
        "ai_summary_mr": None,
        "ai_narrative_en": None,
        "ai_narrative_hi": None,
        "ai_narrative_mr": None,
        "movement_direction": "unknown",
        "movement_confidence": 0.0,
        "narrative_generation_mode": "fallback:not_generated",
    }

    try:
        from .camera import _generate_confirmed_threat_narrative

        narrative_payload = await _generate_confirmed_threat_narrative(
            alert_type=str(alert.get("type") or "unknown"),
            subtype=alert.get("subtype"),
            confidence=float(alert.get("confidence") or 0.0),
            frame_path=selected_frame_path,
            camera_location=location_value,
            incident_time=incident_time_value,
        )
    except Exception as exc:
        print(f"[REPORTS] Narrative generation failed: {exc}")

    now_utc = datetime.now(timezone.utc)
    action = {
        "action": ALERT_STATUS_CONFIRMED,
        "by": email,
        "timestamp": now_utc,
    }

    # Atomic idempotency guard: only one request can claim dispatch for this alert.
    claim_filter = {
        "_id": alert_object_id,
        "$and": [
            {
                "$or": [
                    {"dispatched": {"$exists": False}},
                    {"dispatched": {"$ne": True}},
                ]
            },
            {
                "$or": [
                    {"dispatched_at": {"$exists": False}},
                    {"dispatched_at": None},
                ]
            },
            {"status": {"$nin": [ALERT_STATUS_CONFIRMED, "dispatched"]}},
        ],
    }
    claim_result = await alerts_collection.update_one(
        claim_filter,
        {
            "$set": {
                "verified": True,
                "status": ALERT_STATUS_CONFIRMED,
                "dispatched": True,
                "dispatched_at": now_utc,
                "location": location_value,
            },
            "$push": {"action_history": action},
        },
    )

    if not bool(claim_result.matched_count):
        return {
            "modified_count": 0,
            "message": "Alert already dispatched.",
            "already_dispatched": True,
            "dispatched_recipients": [],
        }

    set_payload = {"location": location_value}
    set_payload.update(narrative_payload)
    if selected_frame_path:
        set_payload["frame_path"] = selected_frame_path
    if primary_camera_id is not None:
        set_payload["primary_camera_id"] = int(primary_camera_id)

    await alerts_collection.update_one(
        {"_id": alert_object_id},
        {
            "$set": set_payload,
        },
    )

    dispatch_results: list[dict[str, Any]] = []
    try:
        users_collection = get_users_collection()
        duty_logs_collection = get_db()["duty_logs"]
        alert_after_confirm = {**alert, **set_payload}
        recipients: list[dict[str, Any]] = []
        cursor = users_collection.find({"role": "guard", "isOnDuty": True})
        async for guard in cursor:
            if not guard.get("phone_number"):
                continue
            preferred_language = normalize_guard_language(
                guard.get("preferred_language")
                or (guard.get("settings") or {}).get("preferred_language")
            )
            guard_id = str(guard.get("_id"))
            recipients.append(
                {
                    "guard_id": guard_id,
                    "full_name": guard.get("full_name"),
                    "email": guard.get("email"),
                    "phone": guard.get("phone_number"),
                    "whatsapp_enabled": bool(guard.get("whatsapp_enabled", True)),
                    "preferred_language": preferred_language,
                }
            )
            await duty_logs_collection.update_many(
                {
                    "guardId": guard_id,
                    "checkOutTime": None,
                },
                {
                    "$inc": {"totalAlertsReceived": 1},
                },
            )

        from .camera import _notify_guard_contacts

        dispatch_results = await _notify_guard_contacts(
            alert_type=alert_after_confirm.get("type"),
            confidence=alert_after_confirm.get("confidence", 0.0),
            frame_path=selected_frame_path,
            subtype=alert_after_confirm.get("subtype"),
            timestamp=alert_after_confirm.get("timestamp").isoformat(
            ) if alert_after_confirm.get("timestamp") else None,
            location=alert_after_confirm.get("location"),
            on_duty_only=True,
            alert_id=str(alert_after_confirm.get("_id")),
            recipients=recipients,
            respect_cooldown=False,
            ai_summary_en=alert_after_confirm.get("ai_summary_en"),
            ai_summary_hi=alert_after_confirm.get("ai_summary_hi"),
            ai_summary_mr=alert_after_confirm.get("ai_summary_mr"),
            ai_narrative_en=alert_after_confirm.get("ai_narrative_en"),
            ai_narrative_hi=alert_after_confirm.get("ai_narrative_hi"),
            ai_narrative_mr=alert_after_confirm.get("ai_narrative_mr"),
            movement_direction=alert_after_confirm.get("movement_direction"),
            movement_confidence=alert_after_confirm.get("movement_confidence"),
            multi_angle_verified=bool(
                alert_after_confirm.get("multi_angle_verified")),
        )
    except Exception as exc:
        print(f"[REPORTS] Notification failed: {exc}")

    # If we know which guard is handling this incident, associate the alert with their active duty record.
    if email:
        try:
            duty_col = get_db()["guard_duty"]
            active = await duty_col.find_one({"email": email, "logout_time": None})
            if active:
                await duty_col.update_one(
                    {"_id": active["_id"]},
                    {"$addToSet": {"alerts_handled": str(alert_id)}},
                )
        except Exception as exc:
            print(f"[REPORTS] Failed to log alert to duty record: {exc}")

    return {
        "modified_count": int(claim_result.modified_count or 0),
        "frame_path_used": selected_frame_path,
        "narrative": {
            "ai_summary_en": set_payload.get("ai_summary_en"),
            "ai_summary_hi": set_payload.get("ai_summary_hi"),
            "ai_summary_mr": set_payload.get("ai_summary_mr"),
            "ai_narrative_en": set_payload.get("ai_narrative_en"),
            "ai_narrative_hi": set_payload.get("ai_narrative_hi"),
            "ai_narrative_mr": set_payload.get("ai_narrative_mr"),
            "movement_direction": set_payload.get("movement_direction"),
            "movement_confidence": set_payload.get("movement_confidence"),
            "narrative_generation_mode": set_payload.get("narrative_generation_mode"),
        },
        "dispatched_recipients": dispatch_results,
    }


@router.get("/analytics/duty-logs")
async def duty_logs(limit: int = 200, include_history: bool = False):
    """Admin audit trail for guard check-in/check-out history."""
    duty_logs_collection = get_db()["duty_logs"]
    users_collection = get_users_collection()

    rows: list[dict[str, Any]] = []
    deduped_rows: dict[str, dict[str, Any]] = {}
    latest_open_login_by_guard: dict[str, datetime] = {}
    now_utc = datetime.now(timezone.utc)
    duty_session_filter = {
        "$and": [
            {"checkInTime": {"$exists": True}},
            {
                "$or": [
                    {"log_type": {"$exists": False}},
                    {"log_type": "session"},
                ]
            },
        ]
    }
    cursor = duty_logs_collection.find(
        duty_session_filter).sort("checkInTime", -1).limit(limit)
    async for doc in cursor:
        guard_id = str(doc.get("guardId") or "")
        guard_profile = None
        if guard_id:
            try:
                from bson import ObjectId
                guard_profile = await users_collection.find_one(
                    {"_id": ObjectId(guard_id)},
                    {"full_name": 1, "phone_number": 1,
                        "isOnDuty": 1, "duty_updated_at": 1},
                )
            except Exception:
                guard_profile = None

        # Default report should reflect live-known guards only; hide orphan rows
        # when guard profile has been deleted.
        if guard_id and not guard_profile and not include_history:
            continue

        check_in_time = _coerce_datetime(doc.get("checkInTime"))
        check_out_time = _coerce_datetime(doc.get("checkOutTime"))
        checkout_synthesized = False

        # Repair stale open rows at read-time so report shows one active row/guard.
        if check_out_time is None and guard_id:
            guard_is_on_duty = bool(
                (guard_profile or {}).get("isOnDuty", False))

            if not guard_is_on_duty:
                check_out_time = _coerce_datetime(
                    (guard_profile or {}).get("duty_updated_at")) or now_utc
                checkout_synthesized = True
            elif guard_id in latest_open_login_by_guard:
                # Older "open" rows are treated as closed when a newer open shift exists.
                check_out_time = latest_open_login_by_guard[guard_id]
                checkout_synthesized = True
            else:
                latest_open_login_by_guard[guard_id] = check_in_time or now_utc

        row = {
            "id": str(doc.get("_id")),
            "guardId": guard_id,
            "guardName": (guard_profile or {}).get("full_name"),
            "phone_number": (guard_profile or {}).get("phone_number"),
            "checkInTime": check_in_time or doc.get("checkInTime"),
            "checkOutTime": check_out_time,
            "totalAlertsReceived": int(doc.get("totalAlertsReceived") or 0),
        }

        check_in_key = (
            check_in_time.isoformat()
            if isinstance(check_in_time, datetime)
            else str(row.get("checkInTime") or "")
        )
        checkout_key = (
            check_out_time.isoformat()
            if isinstance(check_out_time, datetime)
            else ""
        )
        if guard_id and checkout_key:
            dedupe_key = f"closed:{guard_id}:{checkout_key}"
        else:
            dedupe_key = f"{guard_id}:{check_in_key}" if guard_id else str(
                row["id"])

        existing = deduped_rows.get(dedupe_key)
        if not existing:
            deduped_rows[dedupe_key] = row
            continue

        existing_checkin = _coerce_datetime(existing.get("checkInTime"))
        candidate_checkin = _coerce_datetime(row.get("checkInTime"))
        if existing_checkin and candidate_checkin and candidate_checkin < existing_checkin:
            existing["checkInTime"] = row.get("checkInTime")

        existing["totalAlertsReceived"] = max(
            int(existing.get("totalAlertsReceived") or 0),
            int(row.get("totalAlertsReceived") or 0),
        )

        existing_checkout = _coerce_datetime(existing.get("checkOutTime"))
        candidate_checkout = _coerce_datetime(row.get("checkOutTime"))
        if existing_checkout is None and candidate_checkout is not None:
            existing["checkOutTime"] = row.get("checkOutTime")
        elif existing_checkout is not None and candidate_checkout is not None and candidate_checkout > existing_checkout:
            existing["checkOutTime"] = row.get("checkOutTime")

        if not existing.get("guardName") and row.get("guardName"):
            existing["guardName"] = row.get("guardName")
        if not existing.get("phone_number") and row.get("phone_number"):
            existing["phone_number"] = row.get("phone_number")

    # Include WhatsApp ON/OFF duty status events so report view reflects
    # command-based duty changes even when guard uses chat only.
    status_cursor = duty_logs_collection.find(
        {
            "log_type": "status_event",
            "event_type": "whatsapp_duty_status",
        }
    ).sort("timestamp", -1).limit(limit)
    async for doc in status_cursor:
        guard_id = str(doc.get("guardId") or doc.get("guard_id") or "")
        guard_profile = None
        if guard_id:
            try:
                from bson import ObjectId
                guard_profile = await users_collection.find_one(
                    {"_id": ObjectId(guard_id)},
                    {"full_name": 1, "phone_number": 1,
                        "isOnDuty": 1, "duty_updated_at": 1},
                )
            except Exception:
                guard_profile = await users_collection.find_one(
                    {"_id": guard_id},
                    {"full_name": 1, "phone_number": 1,
                        "isOnDuty": 1, "duty_updated_at": 1},
                )

        # Skip orphaned status events that no longer map to a real guard profile.
        if not guard_profile:
            continue

        event_time = _coerce_datetime(doc.get("timestamp")) or now_utc
        action = str(doc.get("action") or "").upper()
        is_on_event = action == "ON_DUTY"
        guard_is_on_duty = bool((guard_profile or {}).get("isOnDuty", False))
        stale_checkout = _coerce_datetime(
            (guard_profile or {}).get("duty_updated_at")) or event_time

        # If ON event is stale and guard is currently OFF duty, show it as closed.
        check_out_for_row = None if (is_on_event and guard_is_on_duty) else (
            event_time if action == "OFF_DUTY" else stale_checkout
        )

        row = {
            "id": str(doc.get("_id")),
            "guardId": guard_id,
            "guardName": (doc.get("name") or (guard_profile or {}).get("full_name")),
            "phone_number": (guard_profile or {}).get("phone_number"),
            "checkInTime": event_time,
            "checkOutTime": check_out_for_row,
            "totalAlertsReceived": 0,
            "activitySource": str(doc.get("source") or "WhatsApp"),
            "activityAction": action or None,
        }

        dedupe_key = f"status:{row['id']}"
        deduped_rows[dedupe_key] = row

    rows = list(deduped_rows.values())
    rows.sort(
        key=lambda item: _coerce_datetime(
            item.get("checkInTime")) or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    if include_history:
        return rows[: max(1, int(limit))]

    # Default view: one latest row per guard to avoid duplicate-looking entries.
    latest_by_guard: list[dict[str, Any]] = []
    seen_guards: set[str] = set()
    for row in rows:
        guard_key = str(row.get("guardId") or "").strip()
        if not guard_key:
            guard_key = _normalize_phone_digits(row.get("phone_number"))
        if not guard_key:
            guard_key = str(row.get("id") or "")

        if guard_key in seen_guards:
            continue

        seen_guards.add(guard_key)
        latest_by_guard.append(row)

    return latest_by_guard[: max(1, int(limit))]


@router.patch("/alerts/{alert_id}/false-alarm")
async def mark_false_alarm(alert_id: str, email: str | None = None):
    """Mark an alert as dismissed."""
    from bson import ObjectId
    from datetime import datetime

    alerts_collection = get_alerts_collection()
    alert = await alerts_collection.find_one({"_id": ObjectId(alert_id)})
    if not alert:
        return {"modified_count": 0}

    action = {
        "action": ALERT_STATUS_DISMISSED,
        "by": email,
        "timestamp": datetime.now(timezone.utc),
    }

    result = await alerts_collection.update_one(
        {"_id": ObjectId(alert_id)},
        {
            "$set": {"verified": False, "status": ALERT_STATUS_DISMISSED},
            "$push": {"action_history": action},
        },
    )

    # Associate this false alarm with the guard's duty record (if available).
    if email:
        try:
            duty_col = get_db()["guard_duty"]
            active = await duty_col.find_one({"email": email, "logout_time": None})
            if active:
                await duty_col.update_one(
                    {"_id": active["_id"]},
                    {"$addToSet": {"alerts_handled": str(alert_id)}},
                )
        except Exception as exc:
            print(f"[REPORTS] Failed to log alert to duty record: {exc}")

    return {"modified_count": result.modified_count}


@router.get("/summary")
async def summary():
    """Get summary statistics of all alerts."""
    alerts_collection = get_alerts_collection()

    total_alerts = await alerts_collection.count_documents({})
    verified_alerts = await alerts_collection.count_documents({"status": {"$in": [ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED]}})
    false_alarms = await alerts_collection.count_documents({"status": ALERT_STATUS_DISMISSED})

    weapon_alerts = await alerts_collection.count_documents({"type": "weapon", "status": {"$in": [ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED]}})
    violence_alerts = await alerts_collection.count_documents({"type": "violence", "status": {"$in": [ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED]}})
    fire_alerts = await alerts_collection.count_documents({"type": "fire", "status": {"$in": [ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED]}})

    return {
        "total_alerts": total_alerts,
        "verified_alerts": verified_alerts,
        "false_alarms": false_alarms,
        "weapon_alerts": weapon_alerts,
        "violence_alerts": violence_alerts,
        "fire_alerts": fire_alerts,
    }


@router.get("/system-runtime-stats")
async def system_runtime_stats():
    """Live dashboard runtime stats for captured frames and storage usage."""
    db = get_db()
    alerts_collection = get_alerts_collection()

    # Captures directory usage (all saved frame files).
    captured_frames, captures_storage_bytes = _directory_stats(CAPTURES_DIR)

    # MongoDB usage for the active database and alerts collection.
    mongo_db_storage_bytes = 0
    mongo_alerts_storage_bytes = 0
    try:
        db_stats = await db.command("dbStats")
        mongo_db_storage_bytes = int(db_stats.get("storageSize") or 0)
    except Exception:
        mongo_db_storage_bytes = 0

    try:
        coll_stats = await db.command("collStats", alerts_collection.name)
        mongo_alerts_storage_bytes = int(coll_stats.get("storageSize") or 0)
    except Exception:
        mongo_alerts_storage_bytes = 0

    return {
        "captured_frames": captured_frames,
        "captures_storage_bytes": captures_storage_bytes,
        "mongodb_storage_bytes": mongo_db_storage_bytes,
        "mongodb_alerts_storage_bytes": mongo_alerts_storage_bytes,
    }


@router.get("/analytics/summary")
async def analytics_summary(date_from: str | None = None, date_to: str | None = None):
    """Professional report summary: verification quality and response metrics."""
    alerts_collection = get_alerts_collection()
    query = _time_filter(date_from, date_to)
    docs = await alerts_collection.find(query).sort("timestamp", -1).to_list(length=None)

    total = len(docs)
    verified_count = 0
    confirmed_count = 0
    two_angle_verified_count = 0
    false_alarm_count = 0
    pending_count = 0
    total_confidence = 0.0
    by_type: dict[str, dict[str, int]] = {}
    response_seconds: list[float] = []
    top_verifiers: dict[str, dict[str, float | int]] = {}

    for doc in docs:
        alert_type = str(doc.get("type") or "unknown").lower()
        by_type.setdefault(alert_type, {
            "total": 0,
            "verified": 0,
            "confirmed": 0,
            "false_alarm": 0,
            "pending": 0,
        })

        by_type[alert_type]["total"] += 1
        status = _normalize_alert_status(doc.get("status"))
        if bool(doc.get("multi_angle_verified")):
            two_angle_verified_count += 1
        if status == ALERT_STATUS_CONFIRMED:
            confirmed_count += 1
            verified_count += 1
            by_type[alert_type]["confirmed"] += 1
            by_type[alert_type]["verified"] += 1
        elif status == ALERT_STATUS_RESOLVED:
            verified_count += 1
            by_type[alert_type]["verified"] += 1
        elif status == ALERT_STATUS_DISMISSED:
            false_alarm_count += 1
            by_type[alert_type]["false_alarm"] += 1
        else:
            pending_count += 1
            by_type[alert_type]["pending"] += 1

        total_confidence += float(doc.get("confidence") or 0.0)

        created_ts = _coerce_datetime(doc.get("timestamp"))
        action_name, action_ts, actor = _extract_resolution(doc)
        if created_ts and action_ts:
            diff = max(0.0, (action_ts - created_ts).total_seconds())
            response_seconds.append(diff)
            if actor:
                stats = top_verifiers.setdefault(
                    str(actor).lower(), {"count": 0, "total_response": 0.0})
                stats["count"] = int(stats["count"]) + 1
                stats["total_response"] = float(stats["total_response"]) + diff

    avg_confidence = round(total_confidence / total, 4) if total else 0.0
    avg_response = round(sum(response_seconds) /
                         len(response_seconds), 2) if response_seconds else 0.0
    sorted_responses = sorted(response_seconds)
    p95 = 0.0
    if sorted_responses:
        idx = min(len(sorted_responses) - 1,
                  max(0, int(len(sorted_responses) * 0.95) - 1))
        p95 = round(sorted_responses[idx], 2)

    top_verifier_rows = []
    for email, stats in top_verifiers.items():
        count = int(stats["count"])
        total_response = float(stats["total_response"])
        top_verifier_rows.append({
            "email": email,
            "handled": count,
            "avg_response_seconds": round(total_response / count, 2) if count else 0.0,
        })
    top_verifier_rows.sort(key=lambda row: row["handled"], reverse=True)

    return {
        "period": {
            "from": date_from,
            "to": date_to,
        },
        "overall": {
            "total_alerts": total,
            "verified_count": verified_count,
            "confirmed_count": confirmed_count,
            "two_angle_verified_count": two_angle_verified_count,
            "false_alarm_count": false_alarm_count,
            "pending_count": pending_count,
            "verified_rate": _safe_ratio(verified_count, total),
            "false_alarm_rate": _safe_ratio(false_alarm_count, total),
            "avg_confidence": avg_confidence,
        },
        "response_metrics": {
            "avg_response_seconds": avg_response,
            "p95_response_seconds": p95,
            "samples": len(response_seconds),
        },
        "by_type": by_type,
        "top_verifiers": top_verifier_rows[:10],
    }


@router.get("/analytics/false-alarms")
async def false_alarm_analytics(limit: int = 50, date_from: str | None = None, date_to: str | None = None):
    """False-alarm quality analysis for Reports view."""
    alerts_collection = get_alerts_collection()
    query = {
        "status": ALERT_STATUS_DISMISSED,
        **_time_filter(date_from, date_to),
    }
    docs = await alerts_collection.find(query).sort("timestamp", -1).to_list(length=None)

    by_type: dict[str, int] = {}
    by_confidence = {
        "low_lt_0_50": 0,
        "medium_0_50_to_0_80": 0,
        "high_gt_0_80": 0,
    }
    by_guard: dict[str, int] = {}

    for doc in docs:
        alert_type = str(doc.get("type") or "unknown").lower()
        by_type[alert_type] = by_type.get(alert_type, 0) + 1

        confidence = float(doc.get("confidence") or 0.0)
        if confidence < 0.50:
            by_confidence["low_lt_0_50"] += 1
        elif confidence <= 0.80:
            by_confidence["medium_0_50_to_0_80"] += 1
        else:
            by_confidence["high_gt_0_80"] += 1

        _, _, actor = _extract_resolution(doc)
        if actor:
            key = str(actor).lower()
            by_guard[key] = by_guard.get(key, 0) + 1

    recent = [_serialize_alert(doc) for doc in docs[: max(1, limit)]]
    top_guards = [{"email": email, "count": count}
                  for email, count in by_guard.items()]
    top_guards.sort(key=lambda row: row["count"], reverse=True)

    return {
        "total": len(docs),
        "by_type": by_type,
        "by_confidence_range": by_confidence,
        "by_guard": top_guards[:10],
        "recent": recent,
    }


@router.get("/analytics/trend")
async def alerts_trend(granularity: str = "daily", date_from: str | None = None, date_to: str | None = None):
    """Time-series trend for total/verified/false alarms."""
    granularity = (granularity or "daily").lower()
    if granularity not in ("hourly", "daily", "weekly"):
        granularity = "daily"

    alerts_collection = get_alerts_collection()
    docs = await alerts_collection.find(_time_filter(date_from, date_to)).sort("timestamp", 1).to_list(length=None)

    buckets: dict[str, dict[str, int]] = {}
    for doc in docs:
        ts = _coerce_datetime(doc.get("timestamp"))
        if ts is None:
            continue
        key = _bucket_label(ts, granularity)
        row = buckets.setdefault(
            key, {"total": 0, "verified": 0, "false_alarm": 0, "pending": 0})
        row["total"] += 1
        status = _normalize_alert_status(doc.get("status"))
        if status in (ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED):
            row["verified"] += 1
        elif status == ALERT_STATUS_DISMISSED:
            row["false_alarm"] += 1
        else:
            row["pending"] += 1

    data = []
    for key in sorted(buckets.keys()):
        row = buckets[key]
        data.append({"bucket": key, **row})

    return {
        "granularity": granularity,
        "period": {
            "from": date_from,
            "to": date_to,
        },
        "data": data,
    }


@router.get("/analytics/guard-performance")
async def guard_performance(date_from: str | None = None, date_to: str | None = None):
    """Guard working-system analysis for Reports section."""
    alerts_collection = get_alerts_collection()
    duty_collection = get_db()["guard_duty"]
    guards_collection = get_users_collection()

    alerts = await alerts_collection.find(_time_filter(date_from, date_to)).to_list(length=None)
    alert_map = {str(doc.get("_id")): doc for doc in alerts}

    guard_docs = await guards_collection.find({"role": "guard", "is_verified": True}).to_list(length=None)
    guard_by_id: dict[str, dict[str, Any]] = {}
    guard_id_by_email: dict[str, str] = {}
    guard_id_by_phone: dict[str, str] = {}

    for guard in guard_docs:
        gid = str(guard.get("_id"))
        guard_by_id[gid] = guard
        email_key = str(guard.get("email") or "").strip().lower()
        phone_key = _normalize_phone_digits(guard.get("phone_number"))
        if email_key:
            guard_id_by_email[email_key] = gid
        if phone_key:
            guard_id_by_phone[phone_key] = gid

    shifts_query = {}
    date_query = _time_filter(date_from, date_to).get("timestamp")
    if date_query:
        # Approximate shift overlap by login time in range for reporting.
        shifts_query = {"login_time": date_query}

    shifts = await duty_collection.find(shifts_query).to_list(length=None)
    # Canonical duty source is guard profile flag. This avoids stale open-shift
    # rows making guards appear active in reports after they clock out.
    on_duty_guard_ids: set[str] = {
        gid for gid, guard in guard_by_id.items() if bool(guard.get("isOnDuty", False))
    }

    def _resolve_guard_for_shift(shift: dict) -> tuple[str | None, dict[str, Any] | None]:
        guard_id = str(shift.get("guard_id") or "").strip()
        if guard_id and guard_id in guard_by_id:
            return guard_id, guard_by_id[guard_id]

        email_key = str(shift.get("email") or "").strip().lower()
        phone_key = _normalize_phone_digits(
            shift.get("phone_number") or shift.get("phone"))

        if phone_key and phone_key in guard_id_by_phone:
            resolved_id = guard_id_by_phone[phone_key]
            return resolved_id, guard_by_id[resolved_id]
        if email_key and email_key in guard_id_by_email:
            resolved_id = guard_id_by_email[email_key]
            return resolved_id, guard_by_id[resolved_id]

        # Skip duty rows that cannot be tied to a real guard account.
        return None, None

    metrics: dict[str, dict[str, Any]] = {}

    for guard_id, guard in guard_by_id.items():
        metrics[guard_id] = {
            "guard_id": guard_id,
            "guard_name": guard.get("full_name") or "Security Personnel",
            "phone_number": guard.get("phone_number"),
            "email": guard.get("email"),
            "duty_status": "on_duty" if guard_id in on_duty_guard_ids else "off_duty",
            "shifts": 0,
            "total_minutes": 0.0,
            "alerts_handled": 0,
            "verified_or_confirmed": 0,
            "false_alarms": 0,
            "avg_response_seconds": 0.0,
            "_response_samples": 0,
            "_response_total": 0.0,
        }

    for shift in shifts:
        key, resolved_guard = _resolve_guard_for_shift(shift)
        if not key or not resolved_guard:
            continue
        row = metrics.setdefault(key, {
            "guard_id": key,
            "guard_name": resolved_guard.get("full_name") or "Security Personnel",
            "phone_number": resolved_guard.get("phone_number"),
            "email": resolved_guard.get("email"),
            "duty_status": "on_duty" if key in on_duty_guard_ids else "off_duty",
            "shifts": 0,
            "total_minutes": 0.0,
            "alerts_handled": 0,
            "verified_or_confirmed": 0,
            "false_alarms": 0,
            "avg_response_seconds": 0.0,
            "_response_samples": 0,
            "_response_total": 0.0,
        })

        row["shifts"] += 1
        login_time = _coerce_datetime(shift.get("login_time"))
        logout_time = _coerce_datetime(
            shift.get("logout_time")) or datetime.now(timezone.utc)
        if login_time:
            row["total_minutes"] += max(0.0, (logout_time -
                                        login_time).total_seconds() / 60.0)

        handled_ids = shift.get("alerts_handled") or []
        if isinstance(handled_ids, list):
            row["alerts_handled"] += len(handled_ids)
            for alert_id in handled_ids:
                doc = alert_map.get(str(alert_id))
                if not doc:
                    continue
                status = _normalize_alert_status(doc.get("status"))
                if status in (ALERT_STATUS_RESOLVED, ALERT_STATUS_CONFIRMED):
                    row["verified_or_confirmed"] += 1
                elif status == ALERT_STATUS_DISMISSED:
                    row["false_alarms"] += 1

                created_ts = _coerce_datetime(doc.get("timestamp"))
                action_name, action_ts, actor = _extract_resolution(doc)
                row_email = str(row.get("email") or "").strip().lower()
                if created_ts and action_ts and actor and row_email and str(actor).lower() == row_email:
                    row["_response_samples"] += 1
                    row["_response_total"] += max(0.0,
                                                  (action_ts - created_ts).total_seconds())

    rows = []
    for row in metrics.values():
        samples = int(row.pop("_response_samples"))
        total_response = float(row.pop("_response_total"))
        row["total_minutes"] = round(float(row["total_minutes"]), 2)
        row["alerts_per_hour"] = round(
            (row["alerts_handled"] / (row["total_minutes"] / 60.0)), 2) if row["total_minutes"] > 0 else 0.0
        row["false_alarm_rate"] = _safe_ratio(
            int(row["false_alarms"]), int(row["alerts_handled"]))
        row["avg_response_seconds"] = round(
            total_response / samples, 2) if samples else 0.0
        rows.append(row)

    rows.sort(key=lambda r: (
        1 if r.get("duty_status") == "on_duty" else 0,
        int(r["alerts_handled"]),
        -float(r["false_alarm_rate"]),
    ), reverse=True)

    on_duty_count = sum(1 for row in rows if row.get(
        "duty_status") == "on_duty")
    return {
        "period": {
            "from": date_from,
            "to": date_to,
        },
        "on_duty_count": on_duty_count,
        "off_duty_count": max(0, len(rows) - on_duty_count),
        "guards": rows,
    }


@router.get("/analytics/admin-activity")
async def admin_activity(limit: int = 200):
    """Admin activity and roster from admin-signup records."""
    users_collection = get_users_collection()
    safe_limit = max(1, min(int(limit), 500))

    docs = await users_collection.find(
        {"role": "admin"},
        {
            "full_name": 1,
            "first_name": 1,
            "middle_name": 1,
            "last_name": 1,
            "email": 1,
            "phone_number": 1,
            "is_verified": 1,
            "session_active": 1,
            "created_at": 1,
            "last_login": 1,
            "last_logout": 1,
        },
    ).sort("created_at", -1).limit(safe_limit).to_list(length=safe_limit)

    rows: list[dict[str, Any]] = []
    verified_count = 0
    active_now_count = 0

    for doc in docs:
        first_name = str(doc.get("first_name") or "").strip()
        middle_name = str(doc.get("middle_name") or "").strip()
        last_name = str(doc.get("last_name") or "").strip()
        full_name = str(doc.get("full_name") or "").strip()
        if not full_name:
            full_name = " ".join(
                part for part in [first_name, middle_name, last_name] if part
            ).strip() or "Administrator"

        created_at = _coerce_datetime(doc.get("created_at"))
        last_login = _coerce_datetime(doc.get("last_login"))
        last_logout = _coerce_datetime(doc.get("last_logout"))
        is_verified = bool(doc.get("is_verified", False))
        session_active = bool(doc.get("session_active", False))

        if is_verified:
            verified_count += 1
        if session_active:
            active_now_count += 1

        rows.append(
            {
                "id": str(doc.get("_id")),
                "full_name": full_name,
                "email": doc.get("email"),
                "phone_number": doc.get("phone_number"),
                "is_verified": is_verified,
                "status": "active" if session_active else "offline",
                "created_at": created_at,
                "last_login": last_login,
                "last_logout": last_logout,
            }
        )

    return {
        "total_admins": len(rows),
        "verified_admins": verified_count,
        "active_admins": active_now_count,
        "admins": rows,
    }


@router.get("/analytics/duty-logs/export")
async def export_duty_logs_excel(days: int = 7, include_admin: bool = True, include_guard: bool = True):
    """Generate an Excel report for guard duty logs and admin activity."""
    if not include_admin and not include_guard:
        raise HTTPException(
            status_code=400, detail="Enable include_admin or include_guard")

    if days not in (3, 7):
        raise HTTPException(status_code=400, detail="days must be 3 or 7")

    now_utc = datetime.now(timezone.utc)
    window_start = now_utc - timedelta(days=days)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if include_guard:
        guard_rows = await duty_logs(limit=5000, include_history=True)
        sheet = workbook.create_sheet("Guard Duty Logs")
        sheet.append(
            [
                "Guard ID",
                "Guard Name",
                "Phone Number",
                "Check In",
                "Check Out",
                "Duration Minutes",
                "Alerts Received",
                "Duty Status",
                "Activity Source",
                "Activity Action",
            ]
        )

        for row in guard_rows:
            check_in_dt = _coerce_datetime(row.get("checkInTime"))
            if check_in_dt is None or check_in_dt < window_start:
                continue

            check_out_dt = _coerce_datetime(row.get("checkOutTime"))
            if check_out_dt is None:
                duration_minutes = max(
                    0.0,
                    (now_utc - check_in_dt).total_seconds() / 60.0,
                )
                duty_status = "on_duty"
            else:
                duration_minutes = max(
                    0.0,
                    (check_out_dt - check_in_dt).total_seconds() / 60.0,
                )
                duty_status = "off_duty"

            sheet.append(
                [
                    str(row.get("guardId") or ""),
                    str(row.get("guardName") or ""),
                    str(row.get("phone_number") or ""),
                    _to_excel_cell_datetime(check_in_dt),
                    _to_excel_cell_datetime(check_out_dt),
                    round(duration_minutes, 2),
                    int(row.get("totalAlertsReceived") or 0),
                    duty_status,
                    str(row.get("activitySource") or "session"),
                    str(row.get("activityAction") or ""),
                ]
            )

    if include_admin:
        admin_payload = await admin_activity(limit=500)
        admins = admin_payload.get("admins") or []
        sheet = workbook.create_sheet("Admin Activity")
        sheet.append(
            [
                "Admin ID",
                "Admin Name",
                "Email",
                "Phone Number",
                "Verification",
                "Current Status",
                "Created At",
                "Last Login",
                "Last Logout",
            ]
        )

        for row in admins:
            created_at = _coerce_datetime(row.get("created_at"))
            if created_at is None or created_at < window_start:
                continue

            sheet.append(
                [
                    str(row.get("id") or ""),
                    str(row.get("full_name") or ""),
                    str(row.get("email") or ""),
                    str(row.get("phone_number") or ""),
                    "verified" if bool(row.get("is_verified")) else "pending",
                    str(row.get("status") or "offline"),
                    _to_excel_cell_datetime(created_at),
                    _to_excel_cell_datetime(row.get("last_login")),
                    _to_excel_cell_datetime(row.get("last_logout")),
                ]
            )

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    export_name = f"duty_report_{days}d_{now_utc.strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_path = EXPORTS_DIR / export_name
    workbook.save(str(export_path))

    return FileResponse(
        path=str(export_path),
        filename=export_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/custom")
async def list_reports(limit: int = 50):
    """Fetch custom reports from reports collection."""
    reports_collection = get_reports_collection()
    cursor = reports_collection.find().sort("created_at", -1).limit(limit)
    reports: List[dict] = []
    async for doc in cursor:
        doc["id"] = str(doc.pop("_id"))
        reports.append(doc)
    return reports
